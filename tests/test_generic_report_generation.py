from dataclasses import replace
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace
import sys
import tempfile
import unittest
from unittest.mock import MagicMock, patch

import pandas as pd
from openpyxl import load_workbook


PROJECT_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_DIR))

from analysis import run_pipeline
from confirmed_relationship_plan import (
    approve_relationship,
    build_approved_join_plan,
    execute_approved_join_plan,
    reject_relationship,
)
from generic_relationship_state import clear_generic_report_if_inputs_changed
from generic_report_generation import (
    CUSTOMER_NOT_PROVIDED_MESSAGE,
    CUSTOMER_PARTIALLY_PROVIDED_MESSAGE,
    INVALID_DATE_ACTION_EXCLUDE_MONTHLY,
    build_data_preparation_summary,
    customer_analysis_available,
    customer_analysis_unavailable_message,
    generate_generic_report,
    return_analysis_available,
    run_generic_report_preflight,
)
from generic_table_reader import read_tabular_sources
from relationship_discovery import discover_relationships
from standard_field_mapping import (
    generate_unified_orders,
    recommend_standard_field_mappings,
    selections_from_recommendations,
)


DATASET_DIR = PROJECT_DIR / "Global+Electronics+Retailer"


def named_csv(name, text):
    source = BytesIO(text.encode("utf-8"))
    source.name = name
    return source


def pizza_original_csv_sources():
    return [
        named_csv(
            "orders.csv",
            "order_id,date,time\n"
            "1,2015-01-01,11:38:36\n"
            "2,2015-01-01,11:57:40\n"
            "3,2015-01-01,12:12:28\n",
        ),
        named_csv(
            "order_details.csv",
            "order_details_id,order_id,pizza_id,quantity\n"
            "1,1,bbq_ckn_l,1\n"
            "2,1,cali_ckn_m,1\n"
            "3,2,bbq_ckn_l,2\n"
            "4,3,thai_ckn_l,1\n",
        ),
        named_csv(
            "pizzas.csv",
            "pizza_id,pizza_type_id,size,price\n"
            "bbq_ckn_l,bbq_ckn,L,20.75\n"
            "cali_ckn_m,cali_ckn,M,16.75\n"
            "thai_ckn_l,thai_ckn,L,20.75\n",
        ),
        named_csv(
            "pizza_types.csv",
            "pizza_type_id,name,category,ingredients\n"
            'bbq_ckn,The Barbecue Chicken Pizza,Chicken,"Chicken, Barbecue Sauce"\n'
            'cali_ckn,The California Chicken Pizza,Chicken,"Chicken, Artichoke"\n'
            'thai_ckn,The Thai Chicken Pizza,Chicken,"Chicken, Pineapple"\n',
        ),
    ]


def base_source_frame():
    return pd.DataFrame(
        {
            "Order Number": ["A-1", "A-2", "A-3", "A-4"],
            "Order Date": ["2024-01-01", "2024-01-02", "2024-02-01", "2024-02-02"],
            "CustomerKey": [10, 11, 12, 13],
            "ProductKey": [100, 101, 102, 103],
            "Unit Price": [10.0, 20.0, 30.0, 40.0],
            "Quantity": [1, 2, 3, 4],
        }
    )


def base_mapping_result():
    frame = base_source_frame()
    recommendations = recommend_standard_field_mappings(frame)
    selections = selections_from_recommendations(recommendations, confirmed=True)
    return generate_unified_orders(frame, selections)


def changed_unified(mapping_result, **columns):
    unified = mapping_result.unified_orders.copy(deep=True)
    for column, values in columns.items():
        unified[column] = values
    return replace(mapping_result, unified_orders=unified)


def empty_plan():
    return SimpleNamespace(fact_table="Sales", fact_table_id="Sales", steps=())


class GenericPreflightRuleTests(unittest.TestCase):
    def test_any_invalid_unit_price_blocks_without_explicit_exclusion(self):
        mapping = changed_unified(base_mapping_result(), unit_price=[10, None, 30, 40])

        preflight = run_generic_report_preflight(mapping)

        self.assertFalse(preflight.can_generate)
        self.assertEqual(preflight.excluded_row_count, 0)
        self.assertTrue(any(issue.issue_code == "invalid_price_or_quantity" for issue in preflight.issues))

    def test_any_invalid_quantity_blocks_without_explicit_exclusion(self):
        mapping = changed_unified(base_mapping_result(), quantity=[1, 0, 3, 4])

        preflight = run_generic_report_preflight(mapping)

        self.assertFalse(preflight.can_generate)
        self.assertEqual(preflight.excluded_row_count, 0)

    def test_invalid_order_id_always_blocks(self):
        mapping = changed_unified(base_mapping_result(), order_id=["A-1", None, "A-3", "A-4"])

        preflight = run_generic_report_preflight(mapping)

        self.assertFalse(preflight.can_generate)
        self.assertTrue(any(issue.issue_code == "invalid_order_id" for issue in preflight.issues))

    def test_discount_rate_outside_zero_to_one_blocks(self):
        mapping = changed_unified(base_mapping_result(), discount_rate=[0, 0, 1.2, 0])

        preflight = run_generic_report_preflight(mapping)

        self.assertFalse(preflight.can_generate)
        self.assertTrue(any(issue.issue_code == "invalid_discount_rate" for issue in preflight.issues))

    def test_invalid_date_requires_explicit_continue_decision(self):
        source = base_source_frame()
        source["Order Date"] = ["2024-01-01", "bad-date", "bad-date", "2024-02-02"]
        mapping = generate_unified_orders(
            source,
            selections_from_recommendations(
                recommend_standard_field_mappings(source), confirmed=True
            ),
        )

        blocked = run_generic_report_preflight(mapping)
        continued = run_generic_report_preflight(
            mapping,
            invalid_date_action=INVALID_DATE_ACTION_EXCLUDE_MONTHLY,
        )

        self.assertFalse(blocked.can_generate)
        self.assertEqual(blocked.monthly_analysis_excluded_row_count, 0)
        self.assertTrue(continued.can_generate)
        self.assertTrue(mapping.success)
        self.assertEqual(continued.monthly_analysis_excluded_row_count, 2)
        self.assertEqual(continued.calculation_row_count, 4)

    def test_rows_are_never_silently_excluded(self):
        mapping = changed_unified(base_mapping_result(), quantity=[1, 0, 3, 4])

        blocked = run_generic_report_preflight(mapping)
        excluded = run_generic_report_preflight(
            mapping,
            exclude_invalid_price_quantity_rows=True,
        )

        self.assertEqual(blocked.excluded_row_count, 0)
        self.assertEqual(blocked.calculation_row_count, 4)
        self.assertTrue(excluded.can_generate)
        self.assertEqual(excluded.excluded_row_count, 1)
        self.assertEqual(excluded.calculation_row_count, 3)
        self.assertIn("quantity", excluded.excluded_rows_detail.iloc[0]["exclusion_reason"])

    def test_mapping_availability_records_assumptions_and_unavailable_fields(self):
        mapping = base_mapping_result()
        availability = {
            item.field_name: item for item in mapping.field_availability
        }

        self.assertEqual(availability["discount_rate"].availability_status, "assumed_default")
        self.assertEqual(availability["discount_rate"].default_value, 0.0)
        self.assertTrue(availability["discount_rate"].user_confirmed)
        self.assertEqual(availability["returned"].availability_status, "not_provided")


class SmallGenericReportTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.mapping = base_mapping_result()
        cls.preflight = run_generic_report_preflight(cls.mapping)
        cls.report_without_expenses = generate_generic_report(
            mapping_result=cls.mapping,
            preflight=cls.preflight,
            discovery_result=None,
            plan=empty_plan(),
            decisions={},
        )
        expenses = BytesIO(
            b"expense_id,date,expense_category,amount\n"
            b"E-1,2024-01-15,Marketing,5\n"
            b"E-2,2024-02-15,Rent,7\n"
        )
        expenses.name = "expenses.csv"
        cls.report_with_expenses = generate_generic_report(
            mapping_result=cls.mapping,
            preflight=cls.preflight,
            discovery_result=None,
            plan=empty_plan(),
            decisions={},
            expenses_source=expenses,
        )

    def test_returned_not_provided_is_not_a_zero_percent_claim(self):
        tables = self.report_without_expenses["report_tables"]

        self.assertFalse(return_analysis_available(tables))
        self.assertTrue(tables["monthly_summary"]["return_rate"].isna().all())
        self.assertFalse(
            (tables["anomalies"]["anomaly_type"] == "return_rate_over_15_percent").any()
        )
        returned_check = tables["post_conversion_data_quality"].loc[
            tables["post_conversion_data_quality"]["check_name"]
            == "invalid_returned_count"
        ].iloc[0]
        self.assertEqual(returned_check["status"], "not_applicable")
        self.assertEqual(int(returned_check["issue_count"]), 0)

    def test_return_chart_availability_gate_is_false(self):
        self.assertFalse(
            return_analysis_available(self.report_without_expenses["report_tables"])
        )

    def test_markdown_marks_return_analysis_not_available(self):
        summary = self.report_without_expenses["summary_text"]

        self.assertIn("Return Analysis: Not Available", summary)
        self.assertIn("Return data was not provided", summary)
        self.assertNotIn("0.0% return", summary)
        self.assertNotIn("treated as not returned", summary)

    def test_no_expenses_skips_finance_module(self):
        finance_status = self.report_without_expenses["report_tables"]["finance_status"]
        status = finance_status.loc[
            finance_status["metric"] == "finance_module_status", "value"
        ].iloc[0]
        self.assertEqual(status, "skipped")

    def test_optional_expenses_generates_finance_module(self):
        tables = self.report_with_expenses["report_tables"]
        status = tables["finance_status"].loc[
            tables["finance_status"]["metric"] == "finance_module_status", "value"
        ].iloc[0]

        self.assertEqual(status, "passed")
        self.assertFalse(tables["cash_flow_summary"].empty)

    def test_excel_and_markdown_download_payloads_are_created(self):
        report = self.report_without_expenses

        self.assertGreater(len(report["excel_bytes"]), 1_000)
        self.assertTrue(report["summary_text"].startswith("# Sales Reporting Summary"))
        workbook = load_workbook(BytesIO(report["excel_bytes"]), read_only=True)
        try:
            self.assertIn("data_preparation_summary", workbook.sheetnames)
            self.assertIn("field_availability", workbook.sheetnames)
            self.assertIn("excluded_rows_detail", workbook.sheetnames)
        finally:
            workbook.close()

    def test_session_signature_prevents_duplicate_generation(self):
        cached_result = object()
        signature = ("mapping", "expenses")
        state = {
            "generic_report_signature": signature,
            "generic_report_result": cached_result,
        }

        unchanged = clear_generic_report_if_inputs_changed(state, signature)
        changed = clear_generic_report_if_inputs_changed(state, ("new",))

        self.assertFalse(unchanged)
        self.assertTrue(changed)
        self.assertNotIn("generic_report_result", state)


class PizzaGenericReportTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.pizza = pd.DataFrame(
            {
                "Order ID": ["P-1", "P-1", "P-2", "P-3"],
                "Order Date": [
                    "2024-01-05",
                    "2024-01-05",
                    "2024-02-10",
                    "2024-03-12",
                ],
                "Product ID": ["MARG", "COLA", "PEPP", "VEG"],
                "Unit Price": [12.0, 3.0, 15.0, 14.0],
                "Quantity": [1, 2, 1, 3],
                "Product Name": [
                    "Margherita",
                    "Cola",
                    "Pepperoni",
                    "Vegetarian",
                ],
                "Category": ["Pizza", "Drink", "Pizza", "Pizza"],
            }
        )
        cls.discovery = discover_relationships({"Pizza Orders": cls.pizza})
        cls.plan = build_approved_join_plan(cls.discovery, "Pizza Orders", [])
        cls.merge_result = execute_approved_join_plan(cls.discovery, cls.plan)
        recommendations = recommend_standard_field_mappings(
            cls.merge_result.merged_frame
        )
        selections = selections_from_recommendations(
            recommendations, confirmed=True
        )
        cls.mapping = generate_unified_orders(
            cls.merge_result.merged_frame, selections
        )
        cls.preflight = run_generic_report_preflight(
            cls.mapping,
            original_fact_row_count=cls.merge_result.fact_row_count,
            merged_row_count=cls.merge_result.final_row_count,
        )
        cls.report = generate_generic_report(
            mapping_result=cls.mapping,
            preflight=cls.preflight,
            discovery_result=cls.discovery,
            plan=cls.plan,
            decisions={},
        )

    def test_pizza_report_succeeds_without_temporary_customer_id(self):
        self.assertTrue(self.mapping.success, self.mapping.errors)
        self.assertTrue(self.mapping.unified_orders["customer_id"].isna().all())
        self.assertFalse(
            self.mapping.unified_orders["customer_id"]
            .astype("string")
            .str.startswith("TEMP-", na=False)
            .any()
        )
        self.assertTrue(self.preflight.can_generate)
        self.assertEqual(self.report["status"], "ready")
        self.assertEqual(len(self.report["report_tables"]["enriched_orders"]), 4)

    def test_pizza_keeps_sales_product_and_category_analysis(self):
        tables = self.report["report_tables"]

        self.assertFalse(tables["monthly_summary"].empty)
        self.assertFalse(tables["category_summary"].empty)
        self.assertFalse(tables["top_products"].empty)
        self.assertEqual(tables["enriched_orders"]["order_id"].nunique(), 3)
        self.assertEqual(tables["enriched_orders"]["quantity"].sum(), 7)

    def test_pizza_skips_customer_analysis_and_top_customer_claim(self):
        tables = self.report["report_tables"]
        summary = self.report["summary_text"]

        self.assertFalse(customer_analysis_available(tables))
        self.assertTrue(tables["customer_summary"].empty)
        self.assertIn(CUSTOMER_NOT_PROVIDED_MESSAGE, summary)
        self.assertNotIn("highest-revenue customer", summary)
        self.assertNotIn("Top Customer", summary)
        customer_checks = tables["validation_report"][
            tables["validation_report"]["check_name"].str.startswith("customer")
        ]
        self.assertFalse(customer_checks.empty)
        self.assertTrue(customer_checks["status"].eq("not_applicable").all())

    def test_pizza_excel_records_customer_unavailable_without_summary_sheet(self):
        workbook = load_workbook(BytesIO(self.report["excel_bytes"]), read_only=True)
        try:
            self.assertNotIn("customer_summary", workbook.sheetnames)
            availability = pd.read_excel(
                BytesIO(self.report["excel_bytes"]),
                sheet_name="field_availability",
            )
            customer = availability.loc[
                availability["field_name"] == "customer_id"
            ].iloc[0]
            self.assertEqual(customer["availability_status"], "not_provided")
            self.assertIn("Customer analysis", customer["notes"])
        finally:
            workbook.close()


class PartialCustomerAvailabilityReportTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        source = base_source_frame()
        source["CustomerKey"] = pd.Series(
            ["C-1", None, "   ", "C-4"], dtype="string"
        )
        recommendations = recommend_standard_field_mappings(source)
        cls.mapping = generate_unified_orders(
            source,
            selections_from_recommendations(recommendations, confirmed=True),
        )
        cls.preflight = run_generic_report_preflight(cls.mapping)
        cls.report = generate_generic_report(
            mapping_result=cls.mapping,
            preflight=cls.preflight,
            discovery_result=None,
            plan=empty_plan(),
            decisions={},
        )

    def test_partial_customer_message_and_counts_are_recorded(self):
        availability = self.report["field_availability"]
        customer = availability.loc[
            availability["field_name"] == "customer_id"
        ].iloc[0]

        self.assertEqual(customer["availability_status"], "partially_provided")
        self.assertEqual(int(customer["provided_row_count"]), 2)
        self.assertEqual(int(customer["total_row_count"]), 4)
        self.assertIn(
            CUSTOMER_PARTIALLY_PROVIDED_MESSAGE,
            self.preflight.warnings,
        )
        self.assertIn(CUSTOMER_PARTIALLY_PROVIDED_MESSAGE, self.report["summary_text"])
        self.assertIn("2 / 4 valid rows", self.report["summary_text"])
        self.assertEqual(
            customer_analysis_unavailable_message(self.report["report_tables"]),
            CUSTOMER_PARTIALLY_PROVIDED_MESSAGE,
        )
        self.assertNotEqual(self.report["status"], "failed")

        excel_availability = pd.read_excel(
            BytesIO(self.report["excel_bytes"]),
            sheet_name="field_availability",
        )
        excel_customer = excel_availability.loc[
            excel_availability["field_name"] == "customer_id"
        ].iloc[0]
        self.assertEqual(int(excel_customer["provided_row_count"]), 2)
        self.assertEqual(int(excel_customer["total_row_count"]), 4)


class PizzaFourCsvEndToEndTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.discovery = discover_relationships(
            read_tabular_sources(pizza_original_csv_sources())
        )
        expected = {
            ("order_details", ("order_id",), "orders", ("order_id",)),
            ("order_details", ("pizza_id",), "pizzas", ("pizza_id",)),
            ("pizzas", ("pizza_type_id",), "pizza_types", ("pizza_type_id",)),
        }
        selected = []
        for candidate in cls.discovery.relationships:
            identity = (
                candidate.left_table,
                candidate.left_columns,
                candidate.right_table,
                candidate.right_columns,
            )
            if identity in expected:
                selected.append(candidate)
        cls.selected_relationships = selected
        decisions = [approve_relationship(candidate) for candidate in selected]
        cls.plan = build_approved_join_plan(
            cls.discovery, "order_details", decisions
        )
        cls.merge_result = execute_approved_join_plan(cls.discovery, cls.plan)

        merged = cls.merge_result.merged_frame
        recommendations = recommend_standard_field_mappings(merged)
        selections = list(
            selections_from_recommendations(recommendations, confirmed=True)
        )
        explicit_sources = {
            "product_id": "pizza_id",
            "unit_price": "pizzas.price",
        }
        selections = [
            replace(
                selection,
                strategy="source",
                source_column=explicit_sources[selection.standard_field],
                confirmed=True,
            )
            if selection.standard_field in explicit_sources
            else selection
            for selection in selections
        ]
        cls.mapping = generate_unified_orders(merged, selections)
        cls.preflight = run_generic_report_preflight(
            cls.mapping,
            original_fact_row_count=cls.merge_result.fact_row_count,
            merged_row_count=cls.merge_result.final_row_count,
        )
        cls.report = generate_generic_report(
            mapping_result=cls.mapping,
            preflight=cls.preflight,
            discovery_result=cls.discovery,
            plan=cls.plan,
            decisions={
                decision.original_candidate_id: decision for decision in decisions
            },
        )

    def test_four_original_csvs_complete_generic_flow(self):
        self.assertEqual(len(self.discovery.tables), 4)
        self.assertEqual(len(self.selected_relationships), 3)
        self.assertEqual(len(self.plan.steps), 3)
        self.assertTrue(self.merge_result.success)
        self.assertEqual(self.merge_result.fact_row_count, 4)
        self.assertEqual(self.merge_result.final_row_count, 4)
        self.assertTrue(self.mapping.success, self.mapping.errors)
        self.assertTrue(self.preflight.can_generate)
        self.assertEqual(self.report["status"], "ready")
        self.assertEqual(len(self.report["report_tables"]["enriched_orders"]), 4)
        customer = self.report["field_availability"].loc[
            self.report["field_availability"]["field_name"] == "customer_id"
        ].iloc[0]
        self.assertEqual(customer["availability_status"], "not_provided")

    def test_customer_conclusions_are_absent_from_all_report_outputs(self):
        tables = self.report["report_tables"]
        self.assertFalse(customer_analysis_available(tables))
        self.assertTrue(tables["customer_summary"].empty)
        self.assertNotIn("highest-revenue customer", self.report["summary_text"])
        self.assertNotIn("Top Customer", self.report["summary_text"])

        customer_checks = tables["validation_report"][
            tables["validation_report"]["check_name"].str.startswith("customer")
        ]
        self.assertTrue(customer_checks["status"].eq("not_applicable").all())
        self.assertTrue(customer_checks["actual_value"].isna().all())

        workbook = load_workbook(BytesIO(self.report["excel_bytes"]), read_only=True)
        try:
            self.assertNotIn("customer_summary", workbook.sheetnames)
            validation = pd.read_excel(
                BytesIO(self.report["excel_bytes"]),
                sheet_name="validation_report",
            )
            excel_customer_checks = validation[
                validation["check_name"].str.startswith("customer")
            ]
            self.assertTrue(
                excel_customer_checks["status"].eq("not_applicable").all()
            )
            self.assertTrue(excel_customer_checks["actual_value"].isna().all())
        finally:
            workbook.close()

    def test_dashboard_kpis_charts_and_insights_skip_customer_conclusions(self):
        import app as dashboard_app

        metric_columns = []

        def make_columns(count):
            columns = [MagicMock() for _ in range(count)]
            metric_columns.extend(columns)
            return columns

        with patch.object(dashboard_app, "st") as streamlit_mock:
            streamlit_mock.columns.side_effect = make_columns
            dashboard_app.show_kpi_cards(self.report["report_tables"])

        metric_labels = [
            call.args[0]
            for column in metric_columns
            for call in column.metric.call_args_list
        ]
        self.assertNotIn("Top Customer", metric_labels)

        with (
            patch.object(dashboard_app, "st") as streamlit_mock,
            patch.object(dashboard_app, "show_monthly_revenue_trend"),
            patch.object(dashboard_app, "show_revenue_by_category"),
            patch.object(dashboard_app, "show_top_products_chart"),
            patch.object(dashboard_app, "show_top_customers_chart") as customer_chart,
            patch.object(dashboard_app, "show_anomalies_by_type_chart"),
        ):
            streamlit_mock.columns.side_effect = lambda count: [
                MagicMock() for _ in range(count)
            ]
            dashboard_app.show_visual_dashboard(self.report)
            customer_chart.assert_not_called()
            info_messages = [
                call.args[0] for call in streamlit_mock.info.call_args_list
            ]
            self.assertIn(CUSTOMER_NOT_PROVIDED_MESSAGE, info_messages)

        with patch.object(dashboard_app, "st") as streamlit_mock:
            dashboard_app.show_business_insights(self.report["report_tables"])
            insight_text = "\n".join(
                str(call.args[0]) for call in streamlit_mock.write.call_args_list
            )
            self.assertNotIn("Top customer", insight_text)

        with (
            patch.object(dashboard_app, "st"),
            patch.object(dashboard_app, "show_table_expander") as expander,
        ):
            dashboard_app.show_detail_tables(self.report)
            expander_titles = [call.args[0] for call in expander.call_args_list]
            self.assertNotIn("Top Customers", expander_titles)


class CustomerNameWithoutIdReportTests(unittest.TestCase):
    def test_customer_name_does_not_enable_customer_analysis(self):
        source = base_source_frame().drop(columns=["CustomerKey"])
        source["Customer Name"] = ["Alice", "Bob", "Alice", "Carol"]
        recommendations = recommend_standard_field_mappings(source)
        mapping = generate_unified_orders(
            source,
            selections_from_recommendations(recommendations, confirmed=True),
        )
        preflight = run_generic_report_preflight(mapping)
        report = generate_generic_report(
            mapping_result=mapping,
            preflight=preflight,
            discovery_result=None,
            plan=empty_plan(),
            decisions={},
        )

        self.assertIn("customer_name", mapping.unified_orders.columns)
        self.assertTrue(mapping.unified_orders["customer_id"].isna().all())
        self.assertFalse(customer_analysis_available(report["report_tables"]))
        self.assertTrue(report["report_tables"]["customer_summary"].empty)
        self.assertNotIn("highest-revenue customer", report["summary_text"])


class MavenGenericEndToEndTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        paths = [
            DATASET_DIR / "Sales.csv",
            DATASET_DIR / "Products.csv",
            DATASET_DIR / "Customers.csv",
            DATASET_DIR / "Stores.csv",
            DATASET_DIR / "Exchange_Rates.csv",
        ]
        cls.discovery = discover_relationships(read_tabular_sources(paths))
        approved = []
        cls.delivery_candidate = None
        expected_keys = {
            "Products": "ProductKey",
            "Customers": "CustomerKey",
            "Stores": "StoreKey",
        }
        for candidate in cls.discovery.relationships:
            mapping = dict(zip(candidate.left_columns, candidate.right_columns))
            known_dimension = (
                candidate.right_table in expected_keys
                and candidate.left_columns == (expected_keys[candidate.right_table],)
            )
            known_exchange = (
                candidate.right_table == "Exchange_Rates"
                and mapping == {"Currency Code": "Currency", "Order Date": "Date"}
            )
            if (known_dimension or known_exchange) and not candidate.blocked:
                approved.append(approve_relationship(candidate))
            elif "Delivery Date" in candidate.left_columns:
                cls.delivery_candidate = candidate
        cls.decisions = {
            item.original_candidate_id: item for item in approved
        }
        if cls.delivery_candidate is not None:
            rejected = reject_relationship(cls.delivery_candidate)
            cls.decisions[rejected.original_candidate_id] = rejected
        cls.plan = build_approved_join_plan(cls.discovery, "Sales", cls.decisions)
        cls.merge_result = execute_approved_join_plan(cls.discovery, cls.plan)
        recommendations = recommend_standard_field_mappings(
            cls.merge_result.merged_frame
        )
        selections = selections_from_recommendations(
            recommendations, confirmed=True
        )
        cls.mapping_result = generate_unified_orders(
            cls.merge_result.merged_frame, selections
        )
        cls.preflight = run_generic_report_preflight(
            cls.mapping_result,
            original_fact_row_count=cls.merge_result.fact_row_count,
            merged_row_count=cls.merge_result.final_row_count,
        )
        cls.report = generate_generic_report(
            mapping_result=cls.mapping_result,
            preflight=cls.preflight,
            discovery_result=cls.discovery,
            plan=cls.plan,
            decisions=cls.decisions,
        )

    def test_maven_generic_mode_generates_end_to_end_report(self):
        self.assertTrue(self.preflight.can_generate)
        self.assertIn(self.report["status"], {"ready", "review_required"})
        self.assertGreater(len(self.report["excel_bytes"]), 100_000)

    def test_maven_enriched_orders_preserves_62884_rows(self):
        enriched = self.report["report_tables"]["enriched_orders"]
        self.assertEqual(len(enriched), 62884)

    def test_maven_generic_merge_uses_composite_rate_key_without_inflation(self):
        exchange_step = next(
            step for step in self.plan.steps if step.right_table == "Exchange_Rates"
        )

        self.assertEqual(
            dict(zip(exchange_step.left_columns, exchange_step.right_columns)),
            {"Currency Code": "Currency", "Order Date": "Date"},
        )
        self.assertEqual(len(exchange_step.left_columns), 2)
        self.assertEqual(self.merge_result.fact_row_count, 62884)
        self.assertEqual(self.merge_result.final_row_count, 62884)
        self.assertTrue(
            all(item.row_growth == 0 for item in self.merge_result.diagnostics)
        )

    def test_maven_customer_analysis_remains_available(self):
        tables = self.report["report_tables"]

        self.assertTrue(customer_analysis_available(tables))
        self.assertEqual(len(tables["customer_summary"]), 11887)
        self.assertIn("highest-revenue customer", self.report["summary_text"])
        workbook = load_workbook(BytesIO(self.report["excel_bytes"]), read_only=True)
        try:
            self.assertIn("customer_summary", workbook.sheetnames)
        finally:
            workbook.close()

    def test_maven_generic_excel_and_markdown_are_available(self):
        self.assertGreater(len(self.report["excel_bytes"]), 100_000)
        self.assertTrue(
            self.report["summary_text"].startswith("# Sales Reporting Summary")
        )
        workbook = load_workbook(BytesIO(self.report["excel_bytes"]), read_only=True)
        try:
            self.assertIn("enriched_orders", workbook.sheetnames)
            self.assertIn("customer_summary", workbook.sheetnames)
            self.assertIn("data_preparation_summary", workbook.sheetnames)
        finally:
            workbook.close()

    def test_confirmed_discount_default_enters_report(self):
        enriched = self.report["report_tables"]["enriched_orders"]
        self.assertTrue(enriched["discount_rate"].eq(0).all())
        availability = self.report["field_availability"]
        status = availability.loc[
            availability["field_name"] == "discount_rate", "availability_status"
        ].iloc[0]
        self.assertEqual(status, "assumed_default")

    def test_return_unavailable_is_blank_in_excel(self):
        workbook = load_workbook(BytesIO(self.report["excel_bytes"]), read_only=True)
        try:
            worksheet = workbook["monthly_summary"]
            headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
            return_column = headers.index("return_rate") + 1
            values = [
                row[0].value
                for row in worksheet.iter_rows(
                    min_row=2,
                    min_col=return_column,
                    max_col=return_column,
                )
            ]
            self.assertTrue(values)
            self.assertTrue(all(value is None for value in values))
        finally:
            workbook.close()

    def test_data_preparation_summary_records_relationships_and_rows(self):
        summary = self.report["data_preparation_summary"]
        approved_rows = summary[summary["section"] == "approved_relationship"]
        rejected_rows = summary[summary["section"] == "rejected_relationship"]
        final_rows = summary.loc[
            summary["item"] == "final_calculation_rows", "value"
        ].iloc[0]

        self.assertEqual(len(approved_rows), 4)
        self.assertGreaterEqual(len(rejected_rows), 1)
        self.assertEqual(int(final_rows), 62884)

class ExistingModeCompatibilityTests(unittest.TestCase):
    def test_single_table_pipeline_still_generates_report(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            temp = Path(temp_dir)
            tables, excel_path, summary_path = run_pipeline(
                PROJECT_DIR / "input" / "sample_orders.csv",
                PROJECT_DIR / "input" / "sample_expenses.csv",
                temp / "single.xlsx",
                temp / "single.md",
            )

            self.assertFalse(tables["enriched_orders"].empty)
            self.assertTrue(excel_path.exists())
            self.assertTrue(summary_path.exists())


if __name__ == "__main__":
    unittest.main()
