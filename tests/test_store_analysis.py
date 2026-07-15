from dataclasses import replace
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace
import sys
import unittest

import pandas as pd
from openpyxl import load_workbook


PROJECT_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_DIR))

from generic_report_generation import (
    generate_generic_report,
    get_field_availability_status,
    run_generic_report_preflight,
)
from generic_relationship_ui import build_single_table_join_result
from generic_table_reader import read_tabular_sources
from generic_store_analysis import (
    STORE_SUMMARY_COLUMNS,
    UNASSIGNED_STORE,
    valid_store_summary,
)
from standard_field_mapping import (
    STORE_NAME_GROUPING_NOTE,
    STORE_NOT_PROVIDED_MESSAGE,
    STORE_PARTIALLY_PROVIDED_MESSAGE,
    build_source_entity_role_map,
    generate_unified_orders,
    recommend_standard_field_mappings,
    selections_from_recommendations,
)
from relationship_discovery import discover_relationships


def sales_frame(**overrides):
    values = {
        "transaction_id": [1, 2, 3, 4],
        "transaction_date": [
            "2024-01-01",
            "2024-01-02",
            "2024-01-03",
            "2024-01-04",
        ],
        "product_id": ["P1", "P1", "P2", "P2"],
        "unit_price": [10.0, 20.0, 5.0, 15.0],
        "transaction_qty": [1, 2, 3, 1],
        "product_detail": ["Coffee", "Coffee", "Tea", "Tea"],
        "product_category": ["Drink", "Drink", "Drink", "Drink"],
    }
    values.update(overrides)
    return pd.DataFrame(values)


def mapping_for(frame, *, source_entity_roles=None, overrides=None):
    recommendations = recommend_standard_field_mappings(
        frame, source_entity_roles
    )
    selections = list(
        selections_from_recommendations(recommendations, confirmed=True)
    )
    for field, (strategy, source) in (overrides or {}).items():
        selections = [
            replace(
                selection,
                strategy=strategy,
                source_column=source,
                confirmed=True,
            )
            if selection.standard_field == field
            else selection
            for selection in selections
        ]
    return generate_unified_orders(
        frame, selections, source_entity_roles=source_entity_roles
    )


def report_for(mapping):
    preflight = run_generic_report_preflight(mapping)
    plan = SimpleNamespace(fact_table="Sales", fact_table_id="Sales", steps=())
    return generate_generic_report(
        mapping_result=mapping,
        preflight=preflight,
        discovery_result=None,
        plan=plan,
        decisions={},
    )


class StoreFieldMappingTests(unittest.TestCase):
    def test_store_aliases_are_recommended_in_english_and_chinese(self):
        english = sales_frame(
            branch_id=[3, 3, 8, 8],
            shop_location=["A", "A", "B", "B"],
        )
        chinese = sales_frame(
            **{
                "门店编号": [3, 3, 8, 8],
                "门店名称": ["甲店", "甲店", "乙店", "乙店"],
            }
        )

        for frame, expected_id, expected_name in (
            (english, "branch_id", "shop_location"),
            (chinese, "门店编号", "门店名称"),
        ):
            recommendations = {
                item.standard_field: item
                for item in recommend_standard_field_mappings(frame)
            }
            self.assertEqual(
                recommendations["store_id"].recommended_source, expected_id
            )
            self.assertEqual(
                recommendations["store_name"].recommended_source, expected_name
            )

    def test_non_store_entities_are_not_recommended_as_store_name(self):
        frame = sales_frame(
            customer_name=["C1", "C2", "C3", "C4"],
            supplier_name=["S1", "S2", "S3", "S4"],
            shipper_name=["H1", "H2", "H3", "H4"],
            employee_name=["E1", "E2", "E3", "E4"],
        )
        roles = {
            "customer_name": "customer",
            "supplier_name": "supplier",
            "shipper_name": "shipper",
            "employee_name": "employee",
        }
        recommendations = {
            item.standard_field: item
            for item in recommend_standard_field_mappings(frame, roles)
        }

        self.assertEqual(
            recommendations["store_name"].recommended_strategy, "omit"
        )

        store_dimension_without_name = sales_frame(
            **{"Stores.Country": ["US", "US", "CA", "CA"]}
        )
        store_recommendations = {
            item.standard_field: item
            for item in recommend_standard_field_mappings(
                store_dimension_without_name,
                {"Stores.Country": "store"},
            )
        }
        self.assertEqual(
            store_recommendations["store_name"].recommended_strategy, "omit"
        )

    def test_store_entity_conflict_disables_store_analysis(self):
        frame = sales_frame(supplier_name=["A", "A", "B", "B"])
        mapping = mapping_for(
            frame,
            source_entity_roles={"supplier_name": "supplier"},
            overrides={"store_name": ("source", "supplier_name")},
        )
        availability = {
            item.field_name: item for item in mapping.field_availability
        }

        self.assertTrue(mapping.success)
        self.assertEqual(
            availability["store_analysis"].availability_status,
            "mapping_conflict",
        )
        report = report_for(mapping)
        self.assertNotIn("store_summary", report["report_tables"])


class StoreAggregationTests(unittest.TestCase):
    def test_provided_id_and_name_use_id_for_grouping_and_reconcile_revenue(self):
        frame = sales_frame(
            store_id=[3, 3, 8, 8],
            store_name=["Astoria", "Astoria renamed", "Manhattan", "Manhattan"],
        )
        mapping = mapping_for(frame)
        report = report_for(mapping)
        summary = report["report_tables"]["store_summary"]
        enriched = report["report_tables"]["enriched_orders"]

        self.assertEqual(
            get_field_availability_status(
                report["report_tables"], "store_analysis"
            ),
            "provided",
        )
        self.assertEqual(tuple(summary.columns), STORE_SUMMARY_COLUMNS)
        self.assertEqual(len(summary), 2)
        self.assertAlmostEqual(
            summary["revenue"].sum(), enriched["final_revenue"].sum()
        )
        self.assertAlmostEqual(summary["orders"].sum(), 4)
        self.assertAlmostEqual(
            summary["units"].sum(), enriched["quantity"].sum()
        )
        store_checks = report["report_tables"]["validation_report"].loc[
            lambda table: table["check_name"].str.startswith("store_summary_")
        ]
        self.assertEqual(set(store_checks["status"]), {"passed"})

    def test_id_only_display_has_no_decimal_suffix(self):
        frame = sales_frame(store_id=[3.0, 3.0, 8.0, 8.0])
        report = report_for(mapping_for(frame))

        self.assertEqual(
            set(report["report_tables"]["store_summary"]["store_name"]),
            {"Store 3", "Store 8"},
        )

    def test_name_only_grouping_is_supported_and_disclosed(self):
        frame = sales_frame(
            store_location=["Astoria", "Astoria", "Soho", "Soho"]
        )
        report = report_for(mapping_for(frame))
        availability = report["field_availability"]
        notes = availability.loc[
            availability["field_name"] == "store_analysis", "notes"
        ].iloc[0]

        self.assertEqual(len(report["report_tables"]["store_summary"]), 2)
        self.assertIn(STORE_NAME_GROUPING_NOTE, notes)

    def test_partial_store_data_uses_unassigned_store_without_losing_revenue(self):
        frame = sales_frame(
            store_id=[3, 3, None, " "],
            store_name=["Astoria", "Astoria", None, ""],
        )
        report = report_for(mapping_for(frame))
        tables = report["report_tables"]
        summary = tables["store_summary"]

        self.assertEqual(
            get_field_availability_status(tables, "store_analysis"),
            "partially_provided",
        )
        self.assertIn(UNASSIGNED_STORE, set(summary["store_name"]))
        self.assertAlmostEqual(
            summary["revenue"].sum(),
            tables["enriched_orders"]["final_revenue"].sum(),
        )
        self.assertNotIn(
            UNASSIGNED_STORE, set(valid_store_summary(summary)["store_name"])
        )
        notes = report["field_availability"].loc[
            report["field_availability"]["field_name"] == "store_analysis",
            "notes",
        ].iloc[0]
        self.assertIn(STORE_PARTIALLY_PROVIDED_MESSAGE, notes)
        self.assertIn("2 of 4 transaction rows", notes)

    def test_store_not_provided_creates_no_module_or_excel_sheet(self):
        report = report_for(mapping_for(sales_frame()))

        self.assertNotIn("store_summary", report["report_tables"])
        self.assertIn(STORE_NOT_PROVIDED_MESSAGE, report["summary_text"])
        workbook = load_workbook(
            BytesIO(report["excel_bytes"]), read_only=True
        )
        try:
            self.assertNotIn("store_summary", workbook.sheetnames)
        finally:
            workbook.close()

    def test_store_summary_is_exported_to_excel_and_markdown(self):
        frame = sales_frame(
            store_id=[3, 3, 8, 8],
            store_location=["Astoria", "Astoria", "Soho", "Soho"],
        )
        report = report_for(mapping_for(frame))
        from app import get_dashboard_metrics

        top = valid_store_summary(
            report["report_tables"]["store_summary"]
        ).iloc[0]

        self.assertEqual(
            get_dashboard_metrics(report["report_tables"])["top_store"],
            top["store_name"],
        )
        self.assertIn("## Store Analysis", report["summary_text"])
        self.assertIn(
            f"Top Store: {top['store_name']}", report["summary_text"]
        )
        workbook = load_workbook(
            BytesIO(report["excel_bytes"]), read_only=True
        )
        try:
            self.assertIn("store_summary", workbook.sheetnames)
            self.assertIn("report_coverage", workbook.sheetnames)
        finally:
            workbook.close()


COFFEE_PATH = PROJECT_DIR / "data" / "raw" / "coffee-shop-sales-revenue.csv"


@unittest.skipUnless(
    COFFEE_PATH.exists(),
    "Full Coffee Shop regression data is not installed in data/raw.",
)
class FullCoffeeStoreEndToEndTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.discovery = discover_relationships(
            read_tabular_sources(COFFEE_PATH)
        )
        cls.plan, cls.merge = build_single_table_join_result(cls.discovery)
        roles = build_source_entity_role_map(
            cls.discovery, cls.plan.fact_table_id
        )
        recommendations = recommend_standard_field_mappings(
            cls.merge.merged_frame, roles
        )
        cls.recommendations = {
            item.standard_field: item for item in recommendations
        }
        selections = [
            replace(
                selection,
                strategy="not_applicable",
                source_column=None,
                confirmed=True,
            )
            if selection.standard_field == "returned"
            else selection
            for selection in selections_from_recommendations(
                recommendations, confirmed=True
            )
        ]
        cls.mapping = generate_unified_orders(
            cls.merge.merged_frame,
            selections,
            source_entity_roles=roles,
        )
        cls.preflight = run_generic_report_preflight(
            cls.mapping,
            original_fact_row_count=cls.merge.fact_row_count,
            merged_row_count=cls.merge.final_row_count,
        )
        cls.report = generate_generic_report(
            mapping_result=cls.mapping,
            preflight=cls.preflight,
            discovery_result=cls.discovery,
            plan=cls.plan,
            decisions={},
        )

    def test_full_coffee_store_report_is_ready_and_reconciled(self):
        self.assertEqual(len(self.discovery.tables), 1)
        self.assertEqual(self.plan.steps, ())
        self.assertEqual(self.merge.final_row_count, 149116)
        self.assertEqual(
            self.recommendations["store_id"].recommended_source, "store_id"
        )
        self.assertEqual(
            self.recommendations["store_name"].recommended_source,
            "store_location",
        )
        self.assertEqual(
            self.recommendations["product_name"].recommended_source,
            "product_detail",
        )
        self.assertEqual(
            self.recommendations["category"].recommended_source,
            "product_category",
        )
        self.assertEqual(self.report["status"], "ready")
        self.assertEqual(len(self.mapping.unified_orders), 149116)
        self.assertEqual(
            get_field_availability_status(
                self.report["report_tables"], "customer_id"
            ),
            "not_provided",
        )
        summary = self.report["report_tables"]["store_summary"]
        enriched = self.report["report_tables"]["enriched_orders"]
        self.assertEqual(len(summary), 3)
        self.assertEqual(summary.iloc[0]["store_name"], "Hell's Kitchen")
        self.assertAlmostEqual(summary["revenue"].sum(), 698812.33, places=2)
        self.assertAlmostEqual(
            summary["revenue"].sum(), enriched["final_revenue"].sum()
        )
        self.assertEqual(int(summary["orders"].sum()), 149116)
        self.assertEqual(int(summary["units"].sum()), 214470)
        self.assertGreater(len(self.report["excel_bytes"]), 1_000)
        self.assertIn("Top Store: Hell's Kitchen", self.report["summary_text"])


if __name__ == "__main__":
    unittest.main()
